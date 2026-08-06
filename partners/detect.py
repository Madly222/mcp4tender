from __future__ import annotations

import openpyxl

from partners.normalize import norm_label

FIELDS = ("brand", "mpn", "description", "cost", "promo", "price_lei", "retail_usd",
          "currency", "warranty", "stock")

SYNONYMS = {
    "brand": ("brand", "producator", "producător", "manufacturer", "marca", "бренд",
              "производитель", "vendor", "make"),
    "mpn": ("p/n", "pn", "part number", "cod", "cod produs", "articol", "sku",
            "артикул", "код", "part no", "mpn", "model"),
    "description": ("description", "descriere", "denumire", "name", "product", "produs",
                    "наименование", "описание", "товар", "denumirea"),
    "cost": ("dealer", "dealer b", "dealer usd", "dealer b, usd", "pret dealer",
             "preț dealer", "cost", "wholesale", "opt", "оптовая", "дилер", "закуп"),
    "promo": ("promo", "promo dealer", "promotie", "promoție", "акция", "promo usd",
              "discount price", "special"),
    "price_lei": ("online, lei", "pret lei", "preț lei", "lei", "mdl", "pret", "preț",
                  "цена lei", "retail lei"),
    "retail_usd": ("online, usd", "retail usd", "pret usd", "preț usd", "цена usd",
                   "online usd", "usd"),
    "currency": ("currency", "valuta", "valută", "moneda", "монеда", "валюта"),
    "warranty": ("warranty", "garantie", "garanție", "гарантия", "garantia"),
    "stock": ("stock", "stoc", "in stock", "disponibil", "наличие", "склад", "qty"),
}

HEADER_HINT = ("brand", "mpn", "description", "cost", "price_lei", "retail_usd")


def _norm_cells(ws, row):
    out = {}
    for c in range(1, ws.max_column + 1):
        lbl = norm_label(ws.cell(row, c).value)
        if lbl:
            out[c] = lbl
    return out


def _match_field(label):
    for field, words in SYNONYMS.items():
        for w in words:
            if label == w:
                return field, 2
    for field, words in SYNONYMS.items():
        for w in words:
            if w in label or label in w:
                return field, 1
    return None, 0


def _score_header(cells):
    hits = 0
    for lbl in cells.values():
        field, strength = _match_field(lbl)
        if field in HEADER_HINT and strength:
            hits += strength
    return hits


def detect_header(ws, scan=30):
    limit = min(scan, ws.max_row)
    best_row, best_score = None, 0
    for r in range(1, limit + 1):
        cells = _norm_cells(ws, r)
        if len(cells) < 2:
            continue
        score = _score_header(cells)
        if score > best_score:
            best_row, best_score = r, score
    if best_row is None:
        return None
    cells = _norm_cells(ws, best_row)
    columns = {}
    used = set()
    for field in FIELDS:
        best_col, best_strength = None, 0
        for col, lbl in cells.items():
            if col in used:
                continue
            f, strength = _match_field(lbl)
            if f == field and strength > best_strength:
                best_col, best_strength = col, strength
        if best_col is not None:
            columns[field] = cells[best_col]
            used.add(best_col)
    return {"header_row": best_row,
            "labels": [cells.get(c, "") for c in range(1, ws.max_column + 1)],
            "columns": columns}


def detect_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        det = detect_header(ws)
        sheets.append({"name": name,
                       "rows": ws.max_row,
                       "images": len(getattr(ws, "_images", [])),
                       "detected": det,
                       "ingest": bool(det and det["columns"].get("mpn")
                                      and (det["columns"].get("cost")
                                           or det["columns"].get("retail_usd")))})
    return {"sheets": sheets}
