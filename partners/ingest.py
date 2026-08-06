from __future__ import annotations

import hashlib
import json
import time

import openpyxl

from partners.normalize import clean_text, norm_label, norm_mpn, to_number
from partners.categorize import classify
from partners.fx import to_usd


def file_sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _header_map(ws, tokens, scan=30):
    want = {norm_label(t) for t in tokens}
    limit = min(scan, ws.max_row)
    for r in range(1, limit + 1):
        labels = {}
        for c in range(1, ws.max_column + 1):
            lbl = norm_label(ws.cell(r, c).value)
            if lbl:
                labels.setdefault(lbl, c)
        if want.issubset(set(labels)):
            return r, labels
    return None, {}


def _resolve_columns(labels, column_spec):
    colmap = {}
    missing = []
    for field, label in column_spec.items():
        col = labels.get(norm_label(label))
        if col is None:
            missing.append(field)
        else:
            colmap[field] = col
    return colmap, missing


def _cell(ws, row, colmap, field):
    col = colmap.get(field)
    if not col:
        return None
    return ws.cell(row, col).value


def _first_nonempty(ws, row):
    for c in range(1, ws.max_column + 1):
        v = clean_text(ws.cell(row, c).value)
        if v:
            return v
    return None


def _outline_level(ws, row):
    dim = ws.row_dimensions.get(row)
    return dim.outline_level if dim is not None else 0


def _classify(ws, row, colmap):
    mpn = clean_text(_cell(ws, row, colmap, "mpn"))
    desc = clean_text(_cell(ws, row, colmap, "description"))
    if mpn or desc:
        return "product"
    if _first_nonempty(ws, row):
        return "category"
    return "blank"


def _crumb_path(crumbs, level):
    parts = [crumbs[k] for k in sorted(crumbs) if k < level and crumbs[k]]
    return " / ".join(parts) if parts else None


def _build_offer(ws, row, colmap, sheet_cfg, crumbs, level, fx_rates):
    cost_field = sheet_cfg.get("cost_field", "cost")
    promo_field = sheet_cfg.get("promo_field", "promo")
    currency = sheet_cfg.get("cost_currency", "USD")
    price_original = to_number(_cell(ws, row, colmap, cost_field)) if cost_field else None
    if price_original is None:
        lei = to_number(_cell(ws, row, colmap, "price_lei"))
        if lei is not None:
            price_original, currency = lei, "MDL"
        else:
            usd = to_number(_cell(ws, row, colmap, "retail_usd"))
            if usd is not None:
                price_original, currency = usd, "USD"
    price_usd = to_usd(price_original, currency, fx_rates)
    promo = to_number(_cell(ws, row, colmap, promo_field)) if promo_field else None
    mpn = clean_text(_cell(ws, row, colmap, "mpn"))
    extra = {}
    for field in ("price_lei", "retail_usd"):
        val = to_number(_cell(ws, row, colmap, field))
        if val is not None:
            extra[field] = val
    description = clean_text(_cell(ws, row, colmap, "description"))
    category_path = _crumb_path(crumbs, level)
    ptype, tsource = classify(description, category_path)
    return {
        "row_no": row,
        "brand": clean_text(_cell(ws, row, colmap, "brand")),
        "mpn": mpn,
        "mpn_norm": norm_mpn(mpn),
        "description": description,
        "category_path": category_path,
        "product_type": ptype,
        "type_source": tsource,
        "price_original": price_original,
        "currency": currency,
        "price_usd": price_usd,
        "promo_usd": promo,
        "warranty": clean_text(_cell(ws, row, colmap, "warranty")),
        "stock": clean_text(_cell(ws, row, colmap, "stock")),
        "extra_json": json.dumps(extra, ensure_ascii=False),
    }


def _ingest_sheet(ws, sheet_cfg, fx_rates):
    header_row, labels = _header_map(ws, sheet_cfg.get("header_tokens", []))
    if header_row is None:
        return [], {"header": "not found", "products": 0, "categories": 0, "unparsed": 0}
    colmap, missing = _resolve_columns(labels, sheet_cfg.get("columns", {}))
    offers = []
    crumbs = {}
    stats = {"header_row": header_row, "products": 0, "categories": 0,
             "unparsed": 0, "missing_columns": missing}
    for row in range(header_row + 1, ws.max_row + 1):
        kind = _classify(ws, row, colmap)
        if kind == "category":
            level = _outline_level(ws, row)
            crumbs[level] = _first_nonempty(ws, row)
            for deeper in [k for k in crumbs if k > level]:
                crumbs.pop(deeper, None)
            stats["categories"] += 1
        elif kind == "product":
            level = _outline_level(ws, row)
            offer = _build_offer(ws, row, colmap, sheet_cfg, crumbs, level, fx_rates)
            if offer["price_original"] is None and not offer["mpn"]:
                stats["unparsed"] += 1
                continue
            offers.append(offer)
            stats["products"] += 1
    return offers, stats


def ingest_workbook(conn, path, profile, fx_rates=None):
    from partners.schema import init_partner_schema
    init_partner_schema(conn)
    partner = profile["partner"]
    sha = file_sha256(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = wb.sheetnames
    report = {"partner": partner, "sheets": {}}
    filename = path.rsplit("/", 1)[-1]
    cur = conn.execute("INSERT INTO partner_files(partner, filename, sha256, received_at, "
                       "sheets_json) VALUES(?,?,?,?,?)",
                       (partner, filename, sha, time.time(), json.dumps(sheets)))
    file_id = cur.lastrowid
    now = time.time()
    superseded = conn.execute(
        "UPDATE partner_offers SET active=0 WHERE partner=? AND active=1",
        (partner,)).rowcount
    report["superseded"] = superseded
    for sheet_cfg in profile["sheets"]:
        name = sheet_cfg["name"]
        if name not in wb.sheetnames:
            report["sheets"][name] = {"skipped": "sheet absent"}
            continue
        if not sheet_cfg.get("ingest"):
            report["sheets"][name] = {"skipped": sheet_cfg.get("reason", "not ingested")}
            continue
        offers, stats = _ingest_sheet(wb[name], sheet_cfg, fx_rates or {})
        for off in offers:
            conn.execute(
                "INSERT INTO partner_offers(file_id, partner, sheet, row_no, brand, mpn, "
                "mpn_norm, description, category_path, product_type, type_source, "
                "price_original, currency, price_usd, "
                "promo_usd, warranty, stock, extra_json, created_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (file_id, partner, name, off["row_no"], off["brand"], off["mpn"],
                 off["mpn_norm"], off["description"], off["category_path"],
                 off["product_type"], off["type_source"],
                 off["price_original"], off["currency"], off["price_usd"], off["promo_usd"],
                 off["warranty"], off["stock"], off["extra_json"], now))
        report["sheets"][name] = stats
    conn.execute("UPDATE partner_files SET report_json=? WHERE id=?",
                 (json.dumps(report, ensure_ascii=False), file_id))
    conn.commit()
    report["file_id"] = file_id
    return report
