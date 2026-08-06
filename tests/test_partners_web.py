from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from fastapi.testclient import TestClient

from engine import accounts, db
from engine.config_store import ConfigStore
from web.server import create_app

HEADER = ["", "BRAND", "P/N", "Description", "Online, LEI", "", "Online, USD",
          "Dealer B, USD", "Promo Dealer, USD ", "", "Warranty", "Stock"]


def _xlsx_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PriceList"
    for c, v in enumerate(HEADER, start=1):
        ws.cell(7, c, v)
    data = [
        ("cat", 0, "PC COMPONENTS"),
        ("cat", 1, "CASE"),
        ("cat", 2, "SHARKOON"),
        ("prod", 3, ("SHARKOON", "RGB SLIDER BK", "Sharkoon RGB SLIDER Black ATX", 890, 48.5, 46, None)),
        ("prod", 3, ("SHARKOON", "RGB FLOW", "Sharkoon RGB FLOW ATX", 1199, 64, 61, None)),
        ("cat", 1, "COOLERS"),
        ("prod", 2, ("DEEPCOOL", "XDC-RF120B", "120mm Case Fan DEEPCOOL", 100, 5.6, 4.7, 4.0)),
    ]
    r = 8
    for kind, lvl, payload in data:
        if kind == "cat":
            ws.cell(r, 1, payload)
        else:
            b, mpn, d, lei, ret, cost, promo = payload
            ws.cell(r, 2, b); ws.cell(r, 3, mpn); ws.cell(r, 4, d)
            ws.cell(r, 5, lei); ws.cell(r, 7, ret); ws.cell(r, 8, cost); ws.cell(r, 9, promo)
        ws.row_dimensions[r].outline_level = lvl
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _client(tmp_path):
    p = str(tmp_path / "w.db")
    conn = db.connect(p); db.init_schema(conn)
    s = ConfigStore(conn); s.reload(); s.seed_defaults("config/defaults")
    accounts.create(conn, "rl", "password1", role="user"); conn.close()
    c = TestClient(create_app(p), follow_redirects=False)
    assert c.post("/login", data={"login": "rl", "password": "password1"}).status_code == 303
    return c


def test_empty_pool_page_renders(tmp_path):
    c = _client(tmp_path)
    h = c.get("/app/partners").text
    assert "Product pool" in h and "Upload price-list" in h


def test_upload_then_confirm_then_pool_shows_products(tmp_path):
    c = _client(tmp_path)
    up = c.post("/app/partners/upload",
                data={"partner": "Accent"},
                files={"file": ("price.xlsx", _xlsx_bytes(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert up.status_code == 303
    loc = up.headers["location"]
    assert "/app/partners/confirm/" in loc
    sid = loc.rsplit("/", 1)[-1]

    conf = c.get(loc).text
    assert "PriceList" in conf and "Article / P/N" in conf

    form = {"partner": "Accent", "sheet_0": "PriceList", "ingest_0": "on",
            "map_0_brand": "brand", "map_0_mpn": "p/n", "map_0_description": "description",
            "map_0_cost": "dealer b, usd", "map_0_promo": "promo dealer, usd",
            "map_0_price_lei": "online, lei", "map_0_retail_usd": "online, usd",
            "map_0_warranty": "warranty", "map_0_stock": "stock", "cur_0": "USD"}
    done = c.post(f"/app/partners/confirm/{sid}", data=form)
    assert done.status_code == 303

    pool = c.get("/app/partners").text
    assert "RGB FLOW" in pool and "XDC-RF120B" in pool
    assert "PC Case" in pool and "Case Fan" in pool


def test_saved_profile_is_reused_and_search_filters(tmp_path):
    c = _client(tmp_path)
    up = c.post("/app/partners/upload", data={"partner": "Accent"},
                files={"file": ("p.xlsx", _xlsx_bytes(), "application/octet-stream")})
    sid = up.headers["location"].rsplit("/", 1)[-1]
    form = {"partner": "Accent", "sheet_0": "PriceList", "ingest_0": "on",
            "map_0_brand": "brand", "map_0_mpn": "p/n", "map_0_description": "description",
            "map_0_cost": "dealer b, usd", "cur_0": "USD"}
    c.post(f"/app/partners/confirm/{sid}", data=form)

    from partners.store import load_profile
    conn = db.connect(str(tmp_path / "w.db"))
    prof = load_profile(conn, "Accent")
    assert prof and prof["sheets"][0]["columns"]["mpn"] == "p/n"

    hit = c.get("/app/partners?q=DEEPCOOL").text
    assert "XDC-RF120B" in hit and "RGB FLOW" not in hit


def test_pool_shows_type_and_filters_by_it(tmp_path):
    c = _client(tmp_path)
    up = c.post("/app/partners/upload", data={"partner": "Accent"},
                files={"file": ("p.xlsx", _xlsx_bytes(), "application/octet-stream")})
    sid = up.headers["location"].rsplit("/", 1)[-1]
    form = {"partner": "Accent", "sheet_0": "PriceList", "ingest_0": "on",
            "map_0_brand": "brand", "map_0_mpn": "p/n", "map_0_description": "description",
            "map_0_cost": "dealer b, usd", "cur_0": "USD"}
    c.post(f"/app/partners/confirm/{sid}", data=form)
    page = c.get("/app/partners").text
    assert "Case Fan" in page and ">Type<" in page
    only = c.get("/app/partners?ptype=Case Fan").text
    assert "XDC-RF120B" in only and "RGB FLOW" not in only
