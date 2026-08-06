from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

from partners.ingest import ingest_workbook
from partners.profiles import get_profile
from partners.schema import init_partner_schema

HEADER = ["", "BRAND", "P/N", "Description", "Online, LEI", "", "Online, USD",
          "Dealer B, USD", "Promo Dealer, USD ", "", "Warranty", "Stock"]


def _make_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PriceList"
    for c, val in enumerate(HEADER, start=1):
        ws.cell(7, c, val)
    rows = [
        ("cat", 0, "PC COMPONENTS"),
        ("cat", 1, "CASE"),
        ("cat", 2, "SHARKOON"),
        ("prod", 3, ("SHARKOON", "RGB SLIDER BK", "Sharkoon RGB SLIDER Black ATX", 890, 48.5, 46, None)),
        ("prod", 3, ("SHARKOON", "RGB FLOW", "Sharkoon RGB FLOW ATX", 1199, 64, 61, None)),
        ("cat", 1, "COOLERS"),
        ("cat", 2, "Case Fan/Fan Control"),
        ("prod", 3, ("DEEPCOOL", "XDC-RF120B", "120mm Case Fan DEEPCOOL", 100, 5.6, 4.7, 4.0)),
        ("cat", 0, "PERIPHERALS"),
        ("cat", 1, "MICE"),
        ("prod", 2, ("LOGITECH", "M90", "Logitech M90 mouse", 120, 6.5, 6.0, None)),
    ]
    r = 8
    for kind, level, payload in rows:
        if kind == "cat":
            ws.cell(r, 1, payload)
        else:
            brand, mpn, desc, lei, retail, dealer, promo = payload
            ws.cell(r, 2, brand)
            ws.cell(r, 3, mpn)
            ws.cell(r, 4, desc)
            ws.cell(r, 5, lei)
            ws.cell(r, 7, retail)
            ws.cell(r, 8, dealer)
            ws.cell(r, 9, promo)
            ws.cell(r, 11, "24 Month")
        ws.row_dimensions[r].outline_level = level
        r += 1
    refurb = wb.create_sheet("Used-Refurb")
    refurb.cell(1, 1, "should be skipped")
    wb.save(path)


def _conn(tmp_path):
    conn = sqlite3.connect(tmp_path / "p.db")
    conn.row_factory = sqlite3.Row
    init_partner_schema(conn)
    return conn


def test_accent_shaped_file_ingests_cleanly(tmp_path):
    f = tmp_path / "accent.xlsx"
    _make_book(f)
    conn = _conn(tmp_path)
    report = ingest_workbook(conn, str(f), get_profile("accent"))
    pl = report["sheets"]["PriceList"]
    assert pl["header_row"] == 7
    assert pl["products"] == 4 and pl["unparsed"] == 0
    assert report["sheets"]["Used-Refurb"]["skipped"]
    n = conn.execute("SELECT COUNT(*) c FROM partner_offers").fetchone()["c"]
    assert n == 4


def test_breadcrumb_follows_outline_levels(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f)
    conn = _conn(tmp_path)
    ingest_workbook(conn, str(f), get_profile("accent"))
    row = conn.execute("SELECT category_path FROM partner_offers WHERE mpn='RGB FLOW'").fetchone()
    assert row["category_path"] == "PC COMPONENTS / CASE / SHARKOON"
    fan = conn.execute("SELECT category_path FROM partner_offers WHERE mpn='XDC-RF120B'").fetchone()
    assert fan["category_path"] == "PC COMPONENTS / COOLERS / Case Fan/Fan Control"
    # a level-2 product under a fresh level-0 category must NOT keep stale deeper crumbs
    mouse = conn.execute("SELECT category_path FROM partner_offers WHERE mpn='M90'").fetchone()
    assert mouse["category_path"] == "PERIPHERALS / MICE"


def test_price_and_promo_are_normalized(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f)
    conn = _conn(tmp_path)
    ingest_workbook(conn, str(f), get_profile("accent"))
    r = conn.execute("SELECT price_usd, promo_usd, currency, mpn_norm, extra_json "
                     "FROM partner_offers WHERE mpn='XDC-RF120B'").fetchone()
    assert r["price_usd"] == 4.7 and r["promo_usd"] == 4.0 and r["currency"] == "USD"
    assert r["mpn_norm"] == "XDC-RF120B"
    assert '"price_lei": 100' in r["extra_json"]


def test_resending_the_same_file_is_rejected(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f)
    conn = _conn(tmp_path)
    ingest_workbook(conn, str(f), get_profile("accent"))
    with pytest.raises(sqlite3.IntegrityError):
        ingest_workbook(conn, str(f), get_profile("accent"))


def test_classify_type_from_keywords_and_path():
    from partners.categorize import classify
    assert classify("MikroTik CRS326 24-port Switch", "NETWORK EQUIPMENT / Switches")[0] == "Network Switch"
    assert classify("8GB DDR3-1600 Kingston", "PC COMPONENTS / RAM / DDR3") == ("RAM", "rule")
    assert classify("APC Back-UPS 650VA", "UPS / Line-Interactive")[0] == "UPS"
    # no rule -> partner top category as a labelled fallback
    t, src = classify("Some obscure gadget", "SMART HOME / Sensors")
    assert t == "Smart Home" and src == "path"
    assert classify(None, None) == (None, None)


def test_offers_get_a_product_type_on_ingest(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f)
    conn = _conn(tmp_path)
    ingest_workbook(conn, str(f), get_profile("accent"))
    r = conn.execute("SELECT product_type, type_source FROM partner_offers "
                     "WHERE mpn='XDC-RF120B'").fetchone()
    assert r["product_type"] == "Case Fan" and r["type_source"] == "rule"
    typed = conn.execute("SELECT COUNT(*) c FROM partner_offers "
                         "WHERE product_type IS NOT NULL").fetchone()["c"]
    assert typed == 4


def test_reclassify_and_type_columns_migrate_onto_old_tables(tmp_path):
    import sqlite3
    from partners.categorize import reclassify_all
    from partners.schema import init_partner_schema
    dbp = tmp_path / "old.db"
    conn = sqlite3.connect(dbp)
    conn.row_factory = sqlite3.Row
    # simulate a pre-stage-5 table without the type columns
    conn.execute("CREATE TABLE partner_offers (id INTEGER PRIMARY KEY AUTOINCREMENT, "
                 "file_id INTEGER, partner TEXT, sheet TEXT, row_no INTEGER, brand TEXT, "
                 "mpn TEXT, mpn_norm TEXT, description TEXT, category_path TEXT, "
                 "price_original REAL, currency TEXT, price_usd REAL, promo_usd REAL, "
                 "warranty TEXT, stock TEXT, extra_json TEXT, created_at REAL)")
    conn.execute("INSERT INTO partner_offers(partner, description, category_path) "
                 "VALUES('X','24-port Switch','NETWORK / Switches')")
    conn.commit()
    init_partner_schema(conn)
    out = reclassify_all(conn)
    assert out["typed"] == 1
    row = conn.execute("SELECT product_type FROM partner_offers").fetchone()
    assert row["product_type"] == "Network Switch"


def test_fx_converts_non_usd_cost_to_usd():
    from partners.fx import to_usd
    fx = {"EUR->MDL": 19.6, "USD->MDL": 18.0, "RON->MDL": 3.9}
    assert to_usd(50, "USD", fx) == 50.0
    assert round(to_usd(100, "EUR", fx), 2) == 108.89
    assert round(to_usd(100, "RON", fx), 2) == 21.67
    assert to_usd(10, "GBP", fx) is None
    assert to_usd(None, "EUR", fx) is None


def test_reingest_supersedes_the_partners_previous_offers(tmp_path):
    import openpyxl
    conn = _conn(tmp_path)
    f1 = tmp_path / "v1.xlsx"
    _make_book(f1)
    ingest_workbook(conn, str(f1), get_profile("accent"))
    active1 = conn.execute("SELECT COUNT(*) c FROM partner_offers WHERE active=1").fetchone()["c"]
    assert active1 == 4

    f2 = tmp_path / "v2.xlsx"
    _make_book(f2)
    wb = openpyxl.load_workbook(f2)
    wb["PriceList"].cell(8, 8, 41)  # change a dealer price so the sha differs
    wb.save(f2)
    ingest_workbook(conn, str(f2), get_profile("accent"))

    active2 = conn.execute("SELECT COUNT(*) c FROM partner_offers WHERE active=1").fetchone()["c"]
    history = conn.execute("SELECT COUNT(*) c FROM partner_offers WHERE active=0").fetchone()["c"]
    assert active2 == 4 and history == 4


def test_non_usd_profile_gets_price_usd_via_fx(tmp_path):
    import openpyxl
    conn = _conn(tmp_path)
    f = tmp_path / "eur.xlsx"
    _make_book(f)
    prof = get_profile("accent")
    prof = {**prof, "sheets": [{**prof["sheets"][0], "cost_currency": "EUR"}] + prof["sheets"][1:]}
    ingest_workbook(conn, str(f), prof, {"EUR->MDL": 19.6, "USD->MDL": 18.0})
    r = conn.execute("SELECT price_original, currency, price_usd FROM partner_offers "
                     "WHERE mpn='XDC-RF120B'").fetchone()
    assert r["currency"] == "EUR" and r["price_original"] == 4.7
    assert round(r["price_usd"], 2) == round(4.7 * 19.6 / 18.0, 2)


def _lei_only_book(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PRODUCTS"
    for c, v in enumerate(["Foto", "Denumire", "Cod Catalog", "Unit.", "Pret, MDL"], start=1):
        ws.cell(4, c, v)
    ws.cell(5, 2, "CABLU CANAL")  # sub-header, no article/price -> dropped
    rows = [("CANAL CABLU TA-GN 60X40", "01780", "m", 124),
            ("CANAL CABLU TA-GN 80X40", "01781", "m", 148.8)]
    r = 6
    for desc, cod, unit, price in rows:
        ws.cell(r, 2, desc); ws.cell(r, 3, cod); ws.cell(r, 4, unit); ws.cell(r, 5, price)
        r += 1
    wb.save(path)


def test_lei_only_file_is_ingestable_and_detected_as_mdl(tmp_path):
    from partners.detect import detect_workbook
    f = tmp_path / "dkc.xlsx"
    _lei_only_book(f)
    det = detect_workbook(str(f))
    sheet = det["sheets"][0]
    assert sheet["ingest"] is True
    assert sheet["detected"]["cost_currency"] == "MDL"
    assert sheet["detected"]["columns"]["cost"] == "pret, mdl"


def test_lei_only_prices_import_and_convert_to_usd(tmp_path):
    f = tmp_path / "dkc.xlsx"
    _lei_only_book(f)
    conn = _conn(tmp_path)
    profile = {"partner": "DKC", "sheets": [{
        "name": "PRODUCTS", "ingest": True,
        "header_tokens": ["cod catalog", "denumire"],
        "columns": {"mpn": "cod catalog", "description": "denumire", "price_lei": "pret, mdl",
                    "cost": "pret, mdl"},
        "cost_field": "cost", "cost_currency": "MDL", "category": "outline"}]}
    ingest_workbook(conn, str(f), profile, {"USD->MDL": 18.0})
    n = conn.execute("SELECT COUNT(*) c FROM partner_offers").fetchone()["c"]
    assert n == 2
    r = conn.execute("SELECT price_original, currency, price_usd FROM partner_offers "
                     "WHERE mpn='01780'").fetchone()
    assert r["price_original"] == 124.0 and r["currency"] == "MDL"
    assert round(r["price_usd"], 2) == round(124.0 / 18.0, 2)
